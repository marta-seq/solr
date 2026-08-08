"""
merge_candidates.py
Reads data/agent_review/staging.xlsx (everything the agents have proposed
so far) and the most recent master Excel in data/data_curated/, applies
candidates into a NEW file - datasets_curated_autoreview_<date>.xlsx -
and NEVER overwrites your original manually-curated file. Archives
staging.xlsx afterward so the next pipeline run starts clean.

Rules:
    - A row with REVIEW_STATUS == "manual" is NEVER touched, full stop -
      belt-and-suspenders on top of triage already excluding these from
      being processed in the first place.
    - If two candidates touch the same cell before you've merged, last-
      written-wins (staging.xlsx is append-only/chronological, so
      processing it top-to-bottom naturally gives this - no special
      tracking needed).
    - needs_review (low-confidence) entries are merged in alongside auto
      entries, same file - REVIEW_STATUS already lets you filter/sort in
      Excel, so a separate file wasn't worth the extra thing to track.
    - As of the method_pub/AP_pub/data schema, every real column
      data_fetch_agent.py/compared_methods_agent.py write (data_DOI,
      paper_DOI, spatial_data_category, spatial_data_method, DOI, category,
      REVIEW_STATUS, ...) has a matching column in "data"/"method_pub"/
      "AP_pub", so nothing needs routing to NEEDS_MANUAL_PLACEMENT anymore -
      that escape hatch existed only because the old Data_multi sheet's
      columns ('DOI '/'DOI' instead of 'data_DOI'/'paper_DOI') didn't line
      up with what the agents actually wrote. Kept as an empty set (rather
      than removed outright) in case a future schema change reintroduces a
      genuinely-mismatched sheet.

Usage:
    python -m src.agents.merge_candidates
"""

import shutil
from datetime import date
from pathlib import Path

import openpyxl

from .common import config

# Header row is NOT the same for every sheet in the current schema:
# method_pub's real header is row 1, but AP_pub and data have a row 1 of
# merged-cell section titles/stray notes with the real header on row 2.
# Rather than hardcode either, scan for the row containing the ID marker -
# same approach 01_parse_excel.py uses, and for the same reason: this is
# exactly the kind of thing that silently breaks when sheets get renamed
# or reshuffled again.
HEADER_SCAN_MAX_ROW = 5
ID_MARKER = "P_ENTRY_ID"


def _find_header_row(ws) -> int:
    for row in range(1, HEADER_SCAN_MAX_ROW + 1):
        for cell in ws[row]:
            if _normalize_header(cell.value) == ID_MARKER:
                return row
    raise ValueError(
        f"Could not find a header row containing '{ID_MARKER}' in the first "
        f"{HEADER_SCAN_MAX_ROW} rows of sheet '{ws.title}'."
    )


# All three sheets use the same ID column now (method_pub used to be the odd
# one out with "PLACEHOLDER_ENTRY_ID" - that's gone since the rename).
ID_COLUMN_BY_SHEET = {
    "method_pub": "P_ENTRY_ID",
    "AP_pub": "P_ENTRY_ID",
    "data": "P_ENTRY_ID",
}

# See module docstring - empty now that "data" unifies the old Data_SP/
# Data_ST/Data_multi columns into one consistent shape.
SHEETS_TOO_AMBIGUOUS_TO_AUTO_MERGE = set()

# staging.xlsx's own bookkeeping columns - never written as a cell value,
# handled separately (as AUTO_* audit columns) instead
_BOOKKEEPING_COLUMNS = (
    "entry_id", "action", "target_sheet", "curation_agent", "curation_model",
    "curation_date", "confidence", "source_paper_entry_id", "notes",
)


def _latest_master_file() -> Path:
    # Excludes "autoreview" outputs (pending review, not current) and Excel's
    # own "~$filename.xlsx" lock/temp files. Deliberately does NOT fall back
    # to "most recently modified" when more than one candidate remains -
    # mtime resets on every `git pull`/clone, so it stops reliably reflecting
    # recency right when you need it most (caught live on gaia: a stale
    # "~$..." lock file looked "newest" and got picked). Fail loudly with a
    # clear fix instead of silently guessing wrong.
    matches = [m for m in config.CURATED_DIR.glob("datasets_curated_*.xlsx")
               if "autoreview" not in m.name.lower() and not m.name.startswith("~$")]
    if not matches:
        raise FileNotFoundError(f"No datasets_curated_*.xlsx found in {config.CURATED_DIR}")
    if len(matches) > 1:
        names = ", ".join(m.name for m in matches)
        raise FileNotFoundError(
            f"Found {len(matches)} candidate master files in {config.CURATED_DIR}, can't tell "
            f"which is current: {names}. Move the stale one(s) into {config.CURATED_DIR.parent / 'data_curated_backup'} "
            f"and leave exactly one in place before re-running."
        )
    return matches[0]


def _normalize_header(h) -> str:
    return str(h).strip() if h is not None else ""


def _header_map(ws, header_row: int) -> dict:
    """{normalized_header_name: column_index} for the real header row."""
    result = {}
    for cell in ws[header_row]:
        name = _normalize_header(cell.value)
        if name:
            result[name] = cell.column
    return result


def _find_row_by_id(ws, id_col_idx: int, entry_id: str, header_row: int):
    for row in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(row=row, column=id_col_idx).value
        if val is not None and str(val).strip() == str(entry_id).strip():
            return row
    return None


def _get_or_add_column(ws, header_map: dict, field_name: str, header_row: int) -> int:
    normalized = _normalize_header(field_name)
    if normalized in header_map:
        return header_map[normalized]
    new_col = ws.max_column + 1
    ws.cell(row=header_row, column=new_col, value=field_name)
    header_map[normalized] = new_col
    return new_col


def _load_staging_records(staging_path: Path) -> list:
    wb = openpyxl.load_workbook(staging_path)
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rec = dict(zip(header, row))
            if rec.get("entry_id"):
                records.append(rec)
    return records


def merge():
    master_path = _latest_master_file()
    print(f"[merge] Using master file: {master_path.name}")

    staging_path = config.STAGING_DIR / "staging.xlsx"
    if not staging_path.exists():
        print("[merge] No staging.xlsx found - nothing to merge.")
        return

    records = _load_staging_records(staging_path)
    if not records:
        print("[merge] staging.xlsx has no candidates - nothing to merge.")
        return

    master_wb = openpyxl.load_workbook(master_path)

    applied = 0
    manual_skipped = 0
    needs_manual_placement = []

    for rec in records:
        action = rec.get("action")
        target_sheet = rec.get("target_sheet")
        entry_id = rec.get("entry_id")
        if not action or not target_sheet or not entry_id:
            continue

        if target_sheet in SHEETS_TOO_AMBIGUOUS_TO_AUTO_MERGE:
            needs_manual_placement.append(rec)
            continue

        if target_sheet not in master_wb.sheetnames:
            print(f"[merge] WARNING: sheet '{target_sheet}' not in master file, skipping {entry_id}")
            continue

        ws = master_wb[target_sheet]
        id_col_name = ID_COLUMN_BY_SHEET.get(target_sheet, "P_ENTRY_ID")
        header_row = _find_header_row(ws)
        header_map = _header_map(ws, header_row)
        id_col_idx = header_map.get(id_col_name)
        if id_col_idx is None:
            print(f"[merge] WARNING: ID column '{id_col_name}' not found in '{target_sheet}', skipping {entry_id}")
            continue

        existing_row = _find_row_by_id(ws, id_col_idx, entry_id, header_row)

        # never touch a manually-reviewed row - belt and suspenders on top
        # of triage already excluding these from being processed at all
        if existing_row is not None:
            review_col = header_map.get("REVIEW_STATUS")
            if review_col:
                current_status = ws.cell(row=existing_row, column=review_col).value
                if str(current_status).strip().lower() == "manual":
                    manual_skipped += 1
                    continue

        candidate_fields = {
            k: v for k, v in rec.items()
            if k not in _BOOKKEEPING_COLUMNS and v not in (None, "")
        }

        if existing_row is None:
            # create_entry (or an update_field that arrived before any
            # create_entry for the same id, which shouldn't normally happen)
            new_row_num = ws.max_row + 1
            id_col_idx = _get_or_add_column(ws, header_map, id_col_name, header_row)
            ws.cell(row=new_row_num, column=id_col_idx, value=entry_id)
            for field_name, value in candidate_fields.items():
                col_idx = _get_or_add_column(ws, header_map, field_name, header_row)
                ws.cell(row=new_row_num, column=col_idx, value=value)
            for audit_field in ("curation_agent", "curation_model", "curation_date", "confidence", "notes"):
                if rec.get(audit_field) not in (None, ""):
                    col_idx = _get_or_add_column(ws, header_map, f"AUTO_{audit_field.upper()}", header_row)
                    ws.cell(row=new_row_num, column=col_idx, value=rec.get(audit_field))
            applied += 1
        else:
            # update_field on an existing row (or create_entry for an id
            # that's somehow already there - treat as an update, don't
            # duplicate the row)
            for field_name, value in candidate_fields.items():
                col_idx = _get_or_add_column(ws, header_map, field_name, header_row)
                ws.cell(row=existing_row, column=col_idx, value=value)
            for audit_field in ("curation_agent", "curation_model", "curation_date", "confidence", "notes"):
                if rec.get(audit_field) not in (None, ""):
                    col_idx = _get_or_add_column(ws, header_map, f"AUTO_{audit_field.upper()}", header_row)
                    ws.cell(row=existing_row, column=col_idx, value=rec.get(audit_field))
            applied += 1

    if needs_manual_placement:
        ws_manual = master_wb.create_sheet("NEEDS_MANUAL_PLACEMENT")
        all_keys = sorted({k for rec in needs_manual_placement for k in rec.keys()})
        ws_manual.append(all_keys)
        for rec in needs_manual_placement:
            ws_manual.append([rec.get(k, "") for k in all_keys])

    today = date.today().isoformat()
    output_path = config.CURATED_DIR / f"datasets_curated_autoreview_{today}.xlsx"
    master_wb.save(output_path)
    print(f"[merge] Wrote {output_path.name}:")
    print(f"[merge]   {applied} candidates applied")
    print(f"[merge]   {manual_skipped} skipped (manually-reviewed rows protected)")
    print(f"[merge]   {len(needs_manual_placement)} routed to NEEDS_MANUAL_PLACEMENT sheet")

    archived_path = config.STAGING_DIR / f"staging_merged_{today}.xlsx"
    shutil.move(str(staging_path), str(archived_path))
    print(f"[merge] Archived staging.xlsx -> {archived_path.name} (next run starts clean)")


if __name__ == "__main__":
    merge()
