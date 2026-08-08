"""
staging.py
Agents NEVER write to the master Excel directly. Instead every proposed
change (new entry, or a new value for an existing entry's field) is saved
into a SEPARATE staging workbook under data/agent_review/, and that workbook
is re-saved to disk after every single candidate - not batched at the end.

ONE ongoing file (data/agent_review/staging.xlsx), not date-stamped. It
keeps accumulating across every run until you actually merge it into the
master Excel - at which point merge_candidates.py should archive/clear it.
Each row's own `curation_date` column already records exactly when that
candidate was found, so the file itself doesn't need to be split by day for
that - and NOT splitting by day is what makes dedup actually work: if it
resumes across a day boundary (very likely given free-tier rate limits),
a fresh run needs to see EVERYTHING not yet merged, not just "today's"
candidates, or it risks creating duplicate entries for the same DOI.

Why an Excel file saved continuously, rather than a JSONL log: free-tier LLM
rate limits (20 req/min, 50-1000/day) mean a run can legitimately die
mid-way through - either killed, or a 429 that exhausts retries. When that
happens you should be able to just open staging.xlsx and see exactly what
was proposed so far, not lose the run or need to parse a log file.

It's still a completely separate file from your master Excel - nothing here
ever touches datasets_curated_*.xlsx directly. A separate merge_candidates.py
script (to be built once this format feels right) will help apply approved
rows into the master file, and should archive staging.xlsx (e.g. rename to
staging_merged_<date>.xlsx) once its contents have been merged, so the next
run starts clean.

Two sheets: "papers" and "datasets" (method_pub/AP_pub both land in "papers",
"data" is the only dataset sheet now - each with a target_sheet column
saying which real sheet it came from, so the merge step routes them).
Columns grow dynamically: whatever keys an agent passes in `fields` become
columns, added on the right the first time they're seen.
"""

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

from . import config

BASE_COLUMNS = [
    "entry_id", "action", "target_sheet", "curation_agent", "curation_model",
    "curation_date", "confidence", "source_paper_entry_id", "notes",
]


def _workbook_path() -> Path:
    return config.STAGING_DIR / "staging.xlsx"


def _load_or_create_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    return wb


def _get_or_create_sheet(wb: Workbook, name: str):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    ws.append(BASE_COLUMNS)
    return ws


def _col_index(ws, key: str) -> int:
    """Returns the 1-indexed column for `key`, creating it (appended on the
    right) if this is the first time this field has been seen."""
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if key in header:
        return header.index(key) + 1
    idx = len(header) + 1
    ws.cell(row=1, column=idx, value=key)
    return idx


def append_candidate(
    action: str,
    sheet: str,
    entry_id: str,
    fields: dict,
    source_paper_entry_id: str,
    curation_agent: str,
    curation_model: str,
    confidence: float = None,
    notes: str = "",
) -> None:
    assert action in ("create_entry", "update_field")
    assert sheet in ("method_pub", "AP_pub", "data")

    path = _workbook_path()
    wb = _load_or_create_workbook(path)
    target = "papers" if sheet in ("method_pub", "AP_pub") else "datasets"
    ws = _get_or_create_sheet(wb, target)

    row_values = {
        "entry_id": entry_id,
        "action": action,
        "target_sheet": sheet,
        "curation_agent": curation_agent,
        "curation_model": curation_model,
        "curation_date": datetime.now(timezone.utc).isoformat(),
        "confidence": confidence,
        "source_paper_entry_id": source_paper_entry_id,
        "notes": notes,
        **fields,
    }

    new_row_num = ws.max_row + 1
    for key, val in row_values.items():
        col = _col_index(ws, key)
        ws.cell(row=new_row_num, column=col, value=val)

    wb.save(path)  # <-- saved to disk immediately, every single candidate


def load_all_candidates_for_run() -> list:
    """Reads back everything staged so far, across every run since the last
    merge - used at startup so a resumed/later run knows what's already been
    proposed and doesn't create duplicate entries before a human has merged
    them into the master DB."""
    path = _workbook_path()
    if not path.exists():
        return []
    wb = load_workbook(path)
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            records.append(dict(zip(header, row)))
    return records
